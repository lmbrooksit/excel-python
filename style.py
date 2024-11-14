from openpyxl.workbook import Workbook # type: ignore
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import Font, Border, Side

# Create a workbook object
wb = load_workbook('colors.xlsx')

# crate an active worksheet
ws = wb.active

# Select Cell
cell = ws['A1']
cell2 = ws['B1']
cell3 = ws['C1']

# Select Cell
cell.font = Font(
    size=30,
    bold=True,
    italic=False,
    color="253bb8"
)

cell2.font = Font(
    size=30,
    bold=False,
    italic=True,
    color="253bb8"
)

cell3.font = Font(
    size=30,
    bold=False,
    italic=False,
    color="253bb8"
)

# Define a Side for our border
my_bd = Side(style="double", color="d80d0d")

B3 = ws['B3']

B3.border = Border(
    left=my_bd,
    right=my_bd,
    top=my_bd,
    bottom=my_bd
)

cell.border = Border(bottom=my_bd)
cell2.border = Border(bottom=my_bd)
cell3.border = Border(bottom=my_bd)


wb.save('names2.xlsx')

print("File Saved")
