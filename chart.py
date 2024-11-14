from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.chart import (
    PieChart,
    Reference,
    Series, 
    PieChart3D,
    BarChart,
    BarChart3D,
    LineChart,
    LineChart3D
)
# load existing spreadsheet

wb = load_workbook('hello.xlsx')

# create an active worksheet
ws = wb.active

# Determine Type of Chart
chart = LineChart()

# Designate Labels and Data
labels = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=10)
data = Reference(ws, min_col=3, min_row=1, max_row=10)

# Put this all together
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

# Add a title
chart.title = "Employee Salaries"

# Place the chart on the spreadsheet
ws.add_chart(chart, "E2")

ws = wb.create_sheet(title="NewSheet")


# Add data in a loop
data = [
    ["Name", "Age"],
    ["Alice", 30],
    ["Bob", 25]
]

for row in data:
    ws.append(row)

# Create new worksheet on the same file
ws = wb.create_sheet(title="NewSheet2")


# Add data in a loop
data = [
    ["Name", "Age"],
    ["Meg", 49],
    ["Jen", 21]
]

for row in data:
    ws.append(row)

wb.save('hello3.xlsx')
print('File Saved')