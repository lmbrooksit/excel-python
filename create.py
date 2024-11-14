from openpyxl.workbook import Workbook # type: ignore
# from openpyxl import load_workbook # type: ignore

# Create a workbook object
wb = Workbook()

# crate an active worksheet
ws = wb.active

# Create worksheet title
ws.title = "Names and Colors"

names = ["Brett", "Taran", "Maltine", "Westly"]
colors = ["white", "black", "tan", "brown"]
nums = [12, 13, 170, 21]

ws['A1'] = "Names"
ws['B1'] = "Colors"
ws['C1'] = "Favorite Number"

# add names to ws
starting_row = 2

for name in names:
    ws.cell(row=starting_row, column=1).value = name
    starting_row += 1

starting_row = 2

for color in colors:
    ws.cell(row=starting_row, column=2).value = color
    starting_row += 1

starting_row = 2

for number in nums:
    ws.cell(row=starting_row, column=3).value = number
    starting_row += 1


ws['C6'] = "=SUM(C2:C5)"
ws['C7'] = "=AVERAGE(C2:C5)"


# Save our spreadsheet
wb.save('colors.xlsx')

print("File Saved")