from openpyxl.workbook import Workbook # type: ignore
from openpyxl import load_workbook # type: ignore

# Create a workbook object
# wb = Workbook()

# load existing spreadsheet
# wb = load_workbook()

# load existing spreadsheet
wb = load_workbook('hello.xlsx')

# create an active worksheet
ws = wb.active

# Set a variable
# name = ws["A3"].value
# color = ws["B3"].value

# Print something from the spreadsheet
# print(f'{name}: {color}')

# Grab a who column
# column_a = ws['7']
# print(column_a)

# For loop
# for cell in column_a:
#     print(cell.value)

# Grab a range 
# range = ws['A2':'B10']

# print(range)

# Loop
# for cell in range:
#     for x in cell:
#         print(x.value)

# Iterate through rows
# for row in ws.iter_rows(min_row=2, max_row=10, min_col=1, max_col=2, values_only=True):
#     for cell in row:
#         print(cell)

# Iterate through columns
# for col in ws.iter_cols(min_row=1, max_row=10, min_col=1, max_col=2, values_only=True):
#     for cell in col:
#         print(cell)

# Create Python List of Names
names = ["Dan", "April", "Neal"]

# Change many cells
starting_row = 12

for name in names:
    ws.cell(row=starting_row, column=1).value = name
    starting_row += 1
# Change many cells

# starting_row = 11
# ws.cell(row=11, column=1).value = "Neo"
# ws.cell(row=11, column=2).value = "Black"

# Change one cell
ws["A2"] = "Johnny"

# Save an excel worksheet
wb.save('hello2.xlsx')

print("file was saved")